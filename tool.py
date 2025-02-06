import os
import pandas as pd
from openpyxl import load_workbook

def get_file_names_without_extension(folder_path):
    """
    フォルダ内のすべてのファイル名（拡張子を除く）を取得します。

    :param folder_path: フォルダのパス
    :return: 拡張子を除いたファイル名のリスト
    """
    file_names = []

    # フォルダを探索
    for root, _, files in os.walk(folder_path):
        for file_name in files:
            # ファイル名と拡張子を分離
            name_without_ext, _ = os.path.splitext(file_name)
            file_names.append(name_without_ext)

    return file_names

def search_elements_in_files(folder_path, elements):
    """
    リスト内の要素がフォルダ内のすべてのファイルの内容に存在するかをチェックします。

    :param folder_path: フォルダのパス
    :param elements: 検索する文字列のリスト
    :return: 要素と対応するファイルを含む辞書
    """
    result = {element: [] for element in elements}  # 結果を格納する辞書

    # フォルダ内のすべてのファイルを探索
    for root, _, files in os.walk(folder_path):
        for file_name in files:
            file_path = os.path.join(root, file_name)
            # ファイル内容を読み取る
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                    # 各要素がファイル内容に存在するかチェック
                    for element in elements:
                        if element in content:
                            result[element].append(file_path)
            except Exception as e:
                print(f"ファイルを読み取れません: {file_path}, エラー: {e}")

    return result


def export_to_excel(data, headers, output_path, sheet_name="Sheet1"):
    """
    データをExcelファイルにエクスポートします。ファイルが存在する場合は新しいシートを追加し、
    存在しない場合は新しいファイルを作成します。

    :param data: データリスト
    :param headers: 列ヘッダーリスト
    :param output_path: Excelファイルのパス
    :param sheet_name: シート名
    """
    # DataFrameを作成
    df = pd.DataFrame(data, columns=headers)

    if os.path.exists(output_path):
        # ファイルが存在する場合、既存のExcelファイルをロード
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="a") as writer:
            # 既存のブックを開く
            book = load_workbook(output_path)

            # シートが存在するかチェック
            if sheet_name in book.sheetnames:
                print(f"シート '{sheet_name}' は既に存在しています。データをそのシートに追加します...")
            else:
                print(f"シート '{sheet_name}' は存在しません。新しいシートを作成します...")

            # データを新しいシートに追加
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            print(f"データが既存ファイル {output_path} に保存されました。シート名: '{sheet_name}'")

    else:
        # ファイルが存在しない場合、新しいファイルを作成
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            print(f"新しいファイルが作成され、{output_path} に保存されました。シート名: '{sheet_name}'")


if __name__=="__main__":

    source_folder_path = r"C:\work\SP"
    destination_folder_path = r"C:\work\物件収支\SRU"
    source_file_list = get_file_names_without_extension(source_folder_path)

    matches = search_elements_in_files(destination_folder_path, source_file_list)

    paris_list = []
    source_extension = os.path.basename(source_folder_path)
    destination_extension = os.path.basename(destination_folder_path)
    headers = [destination_extension, source_extension,]
    sheet_name = destination_extension + "→" + source_extension
    # ID名をリストに格納
    for element, files in matches.items():
        if files:
         for file in files:
          file_name = os.path.splitext(os.path.basename(file))[0]
          paris_list.append([file_name,element])

    export_to_excel(paris_list,headers,r"C:\work\物件収支\物件収支_output.xlsx",sheet_name)
